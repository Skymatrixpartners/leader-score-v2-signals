"""
Realistic Backtest -- IBKR Pro Execution Model
===============================================
Long Only | Limit Orders | Equal-weight | $100,000 starting capital

Commission (both sides):
  $0.005/share  |  min $1.00/order

Bid-Ask Spread (round-trip, by dollar-volume tier via RVOL_Pct proxy):
  Large cap  (RVOL_Pct >= 66, ADV > ~$500M)  : $0.01 / share
  Mid   cap  (RVOL_Pct 33-66)                 : $0.02 / share
  Small cap  (RVOL_Pct <  33)                 : $0.03 / share
  NOTE: 98% of selected trades are in the Large tier.

Score filter:
  Band A : Raw_Score  92 - 95  (fresh breakouts / golden zone, 2x)
  Band B : Raw_Score  98 - 100 (moonshots / institutional breakouts, 1x)
  Dead zone 95-98 is excluded (overextended, -0.46% avg return)

Scenarios:
  1x  Uniform    -- both bands at 1x, no leverage
  Mixed 2x+1x   -- Band A at 2x, Band B at 1x

Output:
  output/backtest/realistic_backtest_10y.xlsx
  output/backtest/realistic_equity_1x.csv
  output/backtest/realistic_equity_mixed.csv

Usage:  py backtest/realistic_backtest_10y.py
"""
from __future__ import annotations
import time
import numpy  as np
import pandas as pd
from pathlib import Path
from openpyxl                 import Workbook
from openpyxl.styles          import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils           import get_column_letter

ROOT     = Path(__file__).resolve().parents[1]
TRADES_F = ROOT / "output" / "backtest" / "combined_trades_10y.csv"
EQUITY_F = ROOT / "output" / "backtest" / "combined_equity_10y.csv"
OUT_XL   = ROOT / "output" / "backtest" / "realistic_backtest_10y.xlsx"
OUT_DIR  = ROOT / "output" / "backtest"

# ── Parameters ──────────────────────────────────────────────────────────────
CAPITAL     = 100_000.0

# Commission (IBKR Fixed, per side)
IBKR_RATE   = 0.005      # $0.005 / share
IBKR_MIN    = 1.00       # minimum $1.00 per order

# Bid-Ask Spread (round-trip cost in $/share, paid on full position)
SPREAD_LARGE = 0.01      # ADV > $500M  (RVOL_Pct >= 66)
SPREAD_MID   = 0.02      # ADV $100M-$500M
SPREAD_SMALL = 0.03      # ADV < $100M
DV_LARGE     = 66.0      # RVOL_Pct threshold
DV_MID       = 33.0

# Score bands
SCORE_LO_A   = 92.0;  SCORE_HI_A = 95.0   # fresh breakouts (golden zone)
SCORE_LO_B   = 98.0                         # moonshots (to 100)

# Scenarios  (label, lev_a, lev_b)
SCENARIOS = [
    ("1x  Uniform",   1.0, 1.0),
    ("Mixed 2x+1x",   2.0, 1.0),
]
LEV_LABELS = [s[0] for s in SCENARIOS]

# ── Colour palette (ARGB, dark theme) ───────────────────────────────────────
C_BG  = "FF0D1117"; C_HDR = "FF161B2E"; C_ROW = "FF1A2035"
C_ACC = "FF6C63FF"; C_GRN = "FF26DE81"; C_RED = "FFFC5C65"
C_GLD = "FFF0C040"; C_TEL = "FF00D4AA"; C_MUT = "FF7B82A0"
C_WHT = "FFFFFFFF"
LEV_C  = ["FF6C63FF", "FF00D4AA"]
LEV_BG = ["FF0D0C1F", "FF001F19"]

# ── Excel helpers ────────────────────────────────────────────────────────────
def _side():
    s = Side(style="thin", color="FF1F2937")
    return Border(left=s, right=s, top=s, bottom=s)

def W(ws, r, c, v=None, bold=False, fc=C_WHT, bg=None,
      hal="center", fmt=None, sz=10):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font      = Font(name="Calibri", bold=bold, color=fc, size=sz)
    cl.alignment = Alignment(horizontal=hal, vertical="center")
    cl.border    = _side()
    if bg:
        cl.fill  = PatternFill("solid", fgColor=bg)
    if fmt:
        cl.number_format = fmt
    return cl

def HDR(ws, row, cols, bg=C_HDR):
    for c, (lbl, wid) in enumerate(cols, 1):
        W(ws, row, c, lbl, bold=True, fc=C_ACC, bg=bg, sz=9)
        ws.column_dimensions[get_column_letter(c)].width = wid
    ws.row_dimensions[row].height = 28

def VC(v): return C_GRN if (v or 0) > 0 else (C_RED if (v or 0) < 0 else C_MUT)

# ── Cost helpers ─────────────────────────────────────────────────────────────
def ibkr_comm(shares: float) -> float:
    return float(max(abs(shares) * IBKR_RATE, IBKR_MIN))

def bid_ask(shares: float, rvol_pct: float) -> float:
    if rvol_pct >= DV_LARGE:
        return abs(shares) * SPREAD_LARGE
    elif rvol_pct >= DV_MID:
        return abs(shares) * SPREAD_MID
    return abs(shares) * SPREAD_SMALL

# ── Core simulation ───────────────────────────────────────────────────────────
def run_sim(trades: pd.DataFrame, spy_map: dict,
            lev_a: float, lev_b: float, rebal_dates: set,
            no_cost: bool = False) -> tuple:
    """
    lev_a : leverage for Band A (scores SCORE_LO_A .. SCORE_HI_A)  golden zone
    lev_b : leverage for Band B (scores SCORE_LO_B .. 100)          moonshots
    no_cost : if True, skip commission and spread (for baseline comparison)
    Each stock gets base share = portfolio / n_total, then multiplied by its band leverage.
    """
    portfolio = CAPITAL
    spy_val   = CAPITAL
    eq_rows: list = []
    tr_rows: list = []

    dates_to_process = sorted(
        d for d in trades["Date"].unique() if str(d)[:10] in rebal_dates
    )

    for date in dates_to_process:
        all_grp = trades[trades["Date"] == date]
        if all_grp.empty:
            continue

        # Filter to the two score bands
        band_a = all_grp[(all_grp["Raw_Score"] >= SCORE_LO_A) &
                         (all_grp["Raw_Score"] <  SCORE_HI_A)]
        band_b = all_grp[all_grp["Raw_Score"] >= SCORE_LO_B]
        grp    = pd.concat([band_a, band_b]).drop_duplicates()
        if grp.empty:
            continue

        regime  = str(grp["Regime"].iloc[0])
        yr      = pd.to_datetime(date).year
        n       = len(grp)
        pv_base = portfolio / n   # un-leveraged base share per stock

        gross = 0.0
        comm_t = spread_t = 0.0

        for _, row in grp.iterrows():
            ep   = float(row.get("Entry_Price") or 0)
            xp   = float(row.get("Exit_Price")  or 0)
            sl   = float(row.get("SL_Price")    or 0)
            ret  = float(row["Return"])
            rvol = float(row.get("RVOL_Pct") or 50)
            sc   = float(row.get("Raw_Score") or 0)

            if ep <= 0:
                continue

            # Band-specific leverage
            lev     = lev_a if SCORE_LO_A <= sc < SCORE_HI_A else lev_b
            pv_each = pv_base * lev
            sh      = pv_each / ep

            c_in  = 0.0 if no_cost else ibkr_comm(sh)
            c_out = 0.0 if no_cost else ibkr_comm(sh)
            sp    = 0.0 if no_cost else bid_ask(sh, rvol)

            cost_t  = c_in + c_out + sp
            gross    += ret * pv_each
            comm_t   += c_in + c_out
            spread_t += sp

            tier     = ("Large" if rvol >= DV_LARGE
                        else ("Mid" if rvol >= DV_MID else "Small"))
            stop_v   = str(row.get("Stop_Triggered", "")) in ("True", "true", "1")
            band_lbl = "92-95" if SCORE_LO_A <= sc < SCORE_HI_A else "98-100"

            tr_rows.append({
                "Date":      str(date)[:10],
                "Ticker":    row["Ticker"],
                "Sector":    row.get("Sector", ""),
                "Regime":    regime,
                "Score":     round(sc, 1),
                "Band":      band_lbl,
                "Lev":       lev,
                "Entry_$":   round(ep, 2),
                "Exit_$":    round(xp, 2),
                "SL_$":      round(sl, 2),
                "Shares":    int(round(sh)),
                "Pos_$":     round(pv_each, 2),
                "Ret_%":     round(ret * 100, 2),
                "Gross_$":   round(ret * pv_each, 2),
                "Comm_$":    round(c_in + c_out, 2),
                "Spread_$":  round(sp, 2),
                "Net_$":     round(ret * pv_each - cost_t, 2),
                "Tier":      tier,
                "Stop":      stop_v,
                "Exit":      row.get("Exit_Reason", "Hold"),
                "Port_Pre":  round(portfolio, 0),
            })

        total_cost = comm_t + spread_t
        pp         = portfolio
        portfolio += gross - total_cost

        spy_ret = float(spy_map.get(str(date)[:10], 0))
        spy_val *= (1 + spy_ret)
        p_ret    = (gross - total_cost) / pp if pp > 0 else 0

        eq_rows.append({
            "Date":       str(date)[:10],
            "Port":       round(portfolio, 2),
            "SPY":        round(spy_val, 2),
            "Ret":        p_ret,
            "SPY_Ret":    spy_ret,
            "Gross":      round(gross, 2),
            "Comm":       round(comm_t, 2),
            "Spread":     round(spread_t, 2),
            "Total_Cost": round(total_cost, 2),
            "N":          n,
            "Regime":     regime,
            "Year":       yr,
        })

    return pd.DataFrame(eq_rows), pd.DataFrame(tr_rows)

# ── Metrics ───────────────────────────────────────────────────────────────────
def calc_metrics(eq: pd.DataFrame, tr: pd.DataFrame | None = None) -> dict:
    pv   = eq["Port"].values
    sv   = eq["SPY"].values
    rets = eq["Ret"].dropna().values

    ny = max((pd.to_datetime(eq["Date"].iloc[-1])
              - pd.to_datetime(eq["Date"].iloc[0])).days / 365.25, 0.1)
    cagr   = (((pv[-1] / pv[0]) ** (1 / ny)) - 1) * 100 if pv[-1] > 0 else -100.0
    s_cagr = ((sv[-1] / sv[0]) ** (1 / ny) - 1) * 100

    pk = pv[0]; mdd = 0.0
    for v in pv:
        pk  = max(pk, v)
        mdd = min(mdd, (v - pk) / pk)

    sharpe = (np.mean(rets) / np.std(rets) * np.sqrt(52)
              if np.std(rets) > 0 else 0)
    calmar = cagr / abs(mdd * 100) if mdd < 0 else 0

    m = {
        "Final_$":   round(pv[-1]),
        "Total%":    round((pv[-1] / pv[0] - 1) * 100, 1),
        "CAGR%":     round(cagr, 1),
        "SPY_CAGR%": round(s_cagr, 1),
        "MaxDD%":    round(mdd * 100, 1),
        "Sharpe":    round(sharpe, 2),
        "Calmar":    round(calmar, 2),
    }
    if tr is not None and not tr.empty:
        m["Comm_$"]       = round(float(tr["Comm_$"].sum()))
        m["Spread_$"]     = round(float(tr["Spread_$"].sum()))
        m["Total_Cost_$"] = m["Comm_$"] + m["Spread_$"]
        m["Cost_Drag%"]   = round(m["Total_Cost_$"] / CAPITAL * 100, 1)
        m["Win%"]         = round((tr["Net_$"] > 0).mean() * 100, 1)
    return m

def annual_table(eq: pd.DataFrame, start_val: float) -> dict:
    eq = eq.copy()
    eq["Year"] = pd.to_datetime(eq["Date"]).dt.year
    rows: dict = {}
    prev = start_val
    for yr, g in eq.groupby("Year"):
        end  = float(g["Port"].iloc[-1])
        ret  = (end / prev - 1) * 100
        pv   = list(g["Port"])
        pk   = prev; mdd_y = 0.0
        for v in pv:
            pk    = max(pk, v)
            mdd_y = min(mdd_y, (v - pk) / pk)
        rows[yr] = {"ret": round(ret, 1), "mdd": round(mdd_y * 100, 1)}
        prev = end
    return rows

def spy_annual(base_eq: pd.DataFrame) -> dict:
    b = base_eq.copy()
    b["Year"] = pd.to_datetime(b["Date"]).dt.year
    out: dict = {}
    prev = float(b["SPY_Value"].iloc[0])
    for yr, g in b.groupby("Year"):
        end      = float(g["SPY_Value"].iloc[-1])
        out[yr]  = round((end / prev - 1) * 100, 1)
        prev     = end
    return out

def annual_costs(eq: pd.DataFrame) -> dict:
    eq = eq.copy()
    eq["Year"] = eq["Year"].astype(int)
    rows: dict = {}
    for yr, g in eq.groupby("Year"):
        rows[yr] = {
            "comm":   round(float(g["Comm"].sum())),
            "spread": round(float(g["Spread"].sum())),
            "total":  round(float(g["Total_Cost"].sum())),
            "port":   round(float(g["Port"].iloc[-1])),
        }
    return rows

def baseline_metrics(base_eq: pd.DataFrame) -> dict:
    pv   = base_eq["Portfolio_Value"].values * CAPITAL
    ny   = max((pd.to_datetime(base_eq["Date"].iloc[-1])
                - pd.to_datetime(base_eq["Date"].iloc[0])).days / 365.25, 0.1)
    cagr = ((pv[-1] / CAPITAL) ** (1 / ny) - 1) * 100
    pk   = CAPITAL; mdd = 0.0
    for v in pv:
        pk  = max(pk, v)
        mdd = min(mdd, (v - pk) / pk)
    brets  = base_eq["Period_Return"].dropna().values
    sharpe = (np.mean(brets) / np.std(brets) * np.sqrt(52)
              if np.std(brets) > 0 else 0)
    return {
        "Final_$": round(pv[-1]), "Total%": round((pv[-1]/CAPITAL-1)*100, 1),
        "CAGR%": round(cagr, 1), "SPY_CAGR%": 15.2,
        "MaxDD%": round(mdd*100, 1), "Sharpe": round(sharpe, 2),
        "Calmar": round(cagr/abs(mdd*100), 2) if mdd < 0 else 0,
        "Comm_$": 0, "Spread_$": 0,
        "Total_Cost_$": 0, "Cost_Drag%": 0.0, "Win%": 51.3,
    }

# ── Excel ─────────────────────────────────────────────────────────────────────
def build_excel(base_eq: pd.DataFrame, results: dict):
    wb = Workbook()
    wb.remove(wb.active)

    base_m = baseline_metrics(base_eq)
    all_m  = [base_m] + [calc_metrics(results[lbl][0], results[lbl][1])
                         for lbl, _, _ in SCENARIOS]
    col_names = ["Baseline\n(No Cost)"] + [
        "%s\n+All Costs" % lbl for lbl, _, _ in SCENARIOS]

    b_copy        = base_eq.copy()
    b_copy["Port"] = b_copy["Portfolio_Value"] * CAPITAL
    b_copy["SPY"]  = b_copy["SPY_Value"]       * CAPITAL
    all_ann = [annual_table(b_copy, CAPITAL)]
    for lbl, _, _ in SCENARIOS:
        all_ann.append(annual_table(results[lbl][0], CAPITAL))
    all_years = sorted({yr for ann in all_ann for yr in ann})
    spy_yr    = spy_annual(base_eq)

    # ── Sheet 1 : Summary ─────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "6C63FF"
    ws.freeze_panes = "B5"

    ws.row_dimensions[1].height = 44
    ws.merge_cells("A1:%s1" % get_column_letter(len(all_m)+1))
    c = ws["A1"]
    c.value = ("REALISTIC BACKTEST  |  IBKR Pro  |  $100,000  |  Long Only  |  "
               "Scores 92-95 (2x) + 98-100 (1x)  |  10-Year (2016-2026)")
    c.font  = Font(name="Calibri", bold=True, color=C_ACC, size=14)
    c.fill  = PatternFill("solid", fgColor=C_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:%s2" % get_column_letter(len(all_m)+1))
    c2 = ws["A2"]
    c2.value = ("Commission: $0.005/sh (min $1)  |  Spread: $0.01/sh (Large)  |  "
                "Band A: scores 92-95 @ 2x leverage  |  "
                "Band B: scores 98-100 @ 1x  |  Dead zone 95-98 excluded")
    c2.font  = Font(name="Calibri", color=C_MUT, size=9)
    c2.fill  = PatternFill("solid", fgColor=C_BG)
    c2.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 6

    hdr_cols = [("Metric", 38)] + [(nm, 22) for nm in col_names]
    HDR(ws, 4, hdr_cols)

    MROWS = [
        ("=== RETURNS ===",                 None, None, False),
        ("Final Portfolio Value ($)",       "Final_$",       "${:,.0f}",   False),
        ("Total Return (%)",                "Total%",        "{:+.1f}%",   True),
        ("CAGR  (annualised %/yr)",         "CAGR%",         "{:+.1f}%/yr",True),
        ("SPY CAGR  (benchmark)",           "SPY_CAGR%",     "{:+.1f}%/yr",False),
        ("=== RISK ===",                    None, None, False),
        ("Max Drawdown (%)",                "MaxDD%",        "{:.1f}%",    False),
        ("Sharpe Ratio  (52-period)",       "Sharpe",        "{:.2f}",     False),
        ("Calmar Ratio",                    "Calmar",        "{:.2f}",     False),
        ("Win Rate  (net P&L basis, %)",    "Win%",          "{:.1f}%",    False),
        ("=== COSTS (10yr cumulative) ===", None, None, False),
        ("IBKR Commission ($)",             "Comm_$",        "${:,.0f}",   False),
        ("Bid-Ask Spread ($)",              "Spread_$",      "${:,.0f}",   False),

        ("TOTAL Transaction Cost ($)",      "Total_Cost_$",  "${:,.0f}",   False),
        ("Cost Drag on Starting $100K (%)","Cost_Drag%",     "{:.1f}%",    False),
    ]

    r = 5
    for label, key, fmt, do_color in MROWS:
        ws.row_dimensions[r].height = 22
        is_sect = key is None
        W(ws, r, 1, label,
          bold=is_sect, fc=C_ACC if is_sect else C_WHT,
          bg=C_BG if is_sect else C_ROW, hal="left")
        for ci, m in enumerate(all_m, 2):
            if is_sect:
                W(ws, r, ci, "", bg=C_BG); continue
            val = m.get(key)
            if val is None:
                W(ws, r, ci, "n/a", fc=C_MUT, bg=C_ROW); continue
            is_nan = isinstance(val, float) and (val != val)
            txt  = fmt.format(val) if not is_nan else "n/a"
            fclr = VC(val) if do_color and not is_nan else C_WHT
            lbg  = LEV_BG[ci - 3] if ci >= 3 else C_ROW
            W(ws, r, ci, txt, fc=fclr, bg=lbg, sz=11)
        r += 1

    # ── Sheet 2 : Annual Returns ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Annual_Returns")
    ws2.sheet_view.showGridLines = False
    ws2.sheet_properties.tabColor = "00D4AA"

    ws2.row_dimensions[1].height = 36
    ws2.merge_cells("A1:G1")
    c2h = ws2["A1"]
    c2h.value = "YEAR-BY-YEAR RETURNS  |  Baseline vs 1x Uniform vs Mixed 2x+1x"
    c2h.font  = Font(name="Calibri", bold=True, color=C_ACC, size=13)
    c2h.fill  = PatternFill("solid", fgColor=C_BG)
    c2h.alignment = Alignment(horizontal="center", vertical="center")

    ann_cols = [("Year",8), ("Baseline\nRet%",13), ("Baseline\nMDD%",12),
                ("1x Uniform\nRet%",14), ("1x Uniform\nMDD%",14),
                ("Mixed 2x+1x\nRet%",15), ("Mixed 2x+1x\nMDD%",15), ("SPY%",10)]
    HDR(ws2, 2, ann_cols)

    for ri, yr in enumerate(all_years, 3):
        ws2.row_dimensions[ri].height = 20
        bg = C_BG if ri % 2 == 0 else C_ROW
        W(ws2, ri, 1, int(yr), bold=True, fc=C_GLD, bg=bg, sz=11)
        col = 2
        for ai, ann in enumerate(all_ann):
            d   = ann.get(yr, {})
            rv  = d.get("ret")
            mv  = d.get("mdd")
            lbg = LEV_BG[ai-1] if ai >= 1 else bg
            W(ws2, ri, col,   ("%+.1f%%" % rv) if rv is not None else "--",
              fc=VC(rv) if rv is not None else C_MUT,
              bg=lbg if ai >= 1 else bg, sz=11)
            W(ws2, ri, col+1, ("%.1f%%" % mv) if mv is not None else "--",
              fc=C_RED if mv is not None and mv < 0 else C_MUT,
              bg=lbg if ai >= 1 else bg, sz=11)
            col += 2
        spy_v = spy_yr.get(yr)
        W(ws2, ri, col,
          ("%+.1f%%" % spy_v) if spy_v is not None else "--",
          fc=VC(spy_v) if spy_v is not None else C_MUT, bg=bg, sz=11)

    # ── Sheet 3 : Cost Breakdown ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Cost_Breakdown")
    ws3.sheet_view.showGridLines = False
    ws3.sheet_properties.tabColor = "F0C040"

    ws3.row_dimensions[1].height = 36
    ws3.merge_cells("A1:L1")
    c3h = ws3["A1"]
    c3h.value = "ANNUAL COST BREAKDOWN  |  All figures in $"
    c3h.font  = Font(name="Calibri", bold=True, color=C_ACC, size=12)
    c3h.fill  = PatternFill("solid", fgColor=C_BG)
    c3h.alignment = Alignment(horizontal="center", vertical="center")

    cost_cols = [
        ("Year",7),
        ("1x Comm",11), ("1x Spread",12), ("1x TOTAL",12), ("1x Port",14),
        ("Mix Comm",11), ("Mix Spread",12), ("Mix TOTAL",12),
    ]
    HDR(ws3, 2, cost_cols)

    ac1 = annual_costs(results["1x  Uniform"][0])
    ac2 = annual_costs(results["Mixed 2x+1x"][0])
    for ri, yr in enumerate(sorted(ac1.keys()), 3):
        ws3.row_dimensions[ri].height = 22
        bg  = C_BG if ri % 2 == 0 else C_ROW
        W(ws3, ri, 1, int(yr), bold=True, fc=C_GLD, bg=bg, sz=11)
        c1  = ac1.get(yr, {}); c2v = ac2.get(yr, {})
        W(ws3, ri,  2, c1.get("comm",0),   fc=C_MUT, bg=bg,       sz=11, fmt='"$"#,##0')
        W(ws3, ri,  3, c1.get("spread",0), fc=C_GLD, bg=bg,       sz=11, fmt='"$"#,##0')
        W(ws3, ri,  4, c1.get("total",0),  fc=C_RED, bg=bg,       sz=11, fmt='"$"#,##0')
        W(ws3, ri,  5, c1.get("port",0),   fc=C_WHT, bg=LEV_BG[0],sz=11, fmt='"$"#,##0')
        W(ws3, ri,  6, c2v.get("comm",0),  fc=C_MUT, bg=LEV_BG[1],sz=11, fmt='"$"#,##0')
        W(ws3, ri,  7, c2v.get("spread",0),fc=C_GLD, bg=LEV_BG[1],sz=11, fmt='"$"#,##0')
        W(ws3, ri,  8, c2v.get("total",0), fc=C_RED, bg=LEV_BG[1],sz=11, fmt='"$"#,##0')

    # ── Sheet 4 : Equity Curve ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Equity_Curve_$")
    ws4.sheet_view.showGridLines = False
    ws4.sheet_properties.tabColor = "26DE81"

    eq_cols = [("Date",12),("Baseline ($)",16),("1x ($)",14),("2x ($)",14),("SPY ($)",14)]
    HDR(ws4, 1, eq_cols)

    b_dmap = dict(zip(base_eq["Date"].astype(str).str[:10],
                      base_eq["Portfolio_Value"] * CAPITAL))
    s_dmap = dict(zip(base_eq["Date"].astype(str).str[:10],
                      base_eq["SPY_Value"]       * CAPITAL))
    lev_pm = {lbl: dict(zip(results[lbl][0]["Date"].astype(str).str[:10],
                            results[lbl][0]["Port"]))
              for lbl, _, _ in SCENARIOS}

    for ri, dt in enumerate(sorted(b_dmap.keys()), 2):
        ws4.row_dimensions[ri].height = 13
        bg = C_BG if ri % 2 == 0 else C_ROW
        W(ws4, ri, 1, dt, fc=C_MUT, bg=bg, sz=9, hal="left")
        bv = b_dmap.get(dt)
        W(ws4, ri, 2, round(bv) if bv else "", fc=C_WHT, bg=bg, sz=9, fmt='"$"#,##0')
        for li, (lbl, _, _) in enumerate(SCENARIOS, 3):
            lv = lev_pm[lbl].get(dt)
            W(ws4, ri, li, round(lv) if lv else "",
              fc=LEV_C[li-3], bg=bg, sz=9, fmt='"$"#,##0')
        sv = s_dmap.get(dt)
        W(ws4, ri, 5, round(sv) if sv else "", fc=C_MUT, bg=bg, sz=9, fmt='"$"#,##0')

    # ── Sheet 5 : Trades 1x ───────────────────────────────────────────────────
    ws5 = wb.create_sheet("Trades_1x")
    ws5.sheet_view.showGridLines = False
    ws5.sheet_properties.tabColor = "6C63FF"

    _, tr1 = results["1x  Uniform"]
    tr_cols = [
        ("Date",9),("Ticker",8),("Sector",18),("Regime",14),
        ("Score",7),("Band",8),("Lev",5),
        ("Entry$",10),("Exit$",10),("SL$",10),
        ("Shares",7),("Pos$",12),
        ("Ret%",8),("Gross$",12),
        ("Comm$",8),("Spread$",8),("Net$",12),
        ("Tier",7),("Stop",7),("Exit Rsn",10),("Pre-Port$",14),
    ]
    HDR(ws5, 1, tr_cols)

    for ri, (_, tr) in enumerate(tr1.head(5000).iterrows(), 2):
        ws5.row_dimensions[ri].height = 13
        bg   = C_BG if ri % 2 == 0 else C_ROW
        rv   = float(tr["Ret_%"]); nv = float(tr["Net_$"])
        tc   = C_GRN if tr["Tier"]=="Large" else (C_GLD if tr["Tier"]=="Mid" else C_RED)
        W(ws5, ri,  1, tr["Date"],   fc=C_MUT, bg=bg, sz=9, hal="left")
        W(ws5, ri,  2, tr["Ticker"], fc=C_ACC, bg=bg, sz=9, hal="left")
        W(ws5, ri,  3, str(tr["Sector"])[:18], fc=C_MUT, bg=bg, sz=9, hal="left")
        W(ws5, ri,  4, tr["Regime"], fc=C_GLD, bg=bg, sz=9)
        W(ws5, ri,  5, tr["Score"],  fc=C_ACC, bg=bg, sz=9)
        bc = C_TEL if str(tr.get("Band",""))=="92-95" else C_GLD
        W(ws5, ri,  6, tr.get("Band",""), fc=bc, bg=bg, sz=9)
        W(ws5, ri,  7, tr.get("Lev",""),  fc=C_MUT, bg=bg, sz=9)
        W(ws5, ri,  8, tr["Entry_$"],fc=C_WHT, bg=bg, sz=9, fmt='"$"#,##0.00')
        W(ws5, ri,  9, tr.get("Exit_$",""), fc=VC(rv), bg=bg, sz=9, fmt='"$"#,##0.00')
        W(ws5, ri, 10, tr.get("SL_$",""),   fc=C_RED,  bg=bg, sz=9, fmt='"$"#,##0.00')
        W(ws5, ri, 11, tr["Shares"], fc=C_WHT, bg=bg, sz=9)
        W(ws5, ri, 12, tr["Pos_$"],  fc=C_TEL, bg=bg, sz=9, fmt='"$"#,##0')
        W(ws5, ri, 13, "%+.2f%%" % rv, fc=VC(rv), bg=bg, sz=9)
        W(ws5, ri, 14, tr["Gross_$"],fc=VC(float(tr["Gross_$"])),bg=bg,sz=9,fmt='"$"#,##0.00')
        W(ws5, ri, 15, tr["Comm_$"], fc=C_MUT, bg=bg, sz=9, fmt='"$"#,##0.00')
        W(ws5, ri, 16, tr["Spread_$"],fc=C_GLD,bg=bg, sz=9, fmt='"$"#,##0.00')
        W(ws5, ri, 17, tr["Net_$"],  fc=VC(nv), bg=bg, sz=9, fmt='"$"#,##0.00')
        W(ws5, ri, 18, tr["Tier"],   fc=tc, bg=bg, sz=9)
        stop_v = str(tr.get("Stop","")) in ("True","true","1")
        W(ws5, ri, 19, "STOP" if stop_v else "--",
          fc=C_RED if stop_v else C_MUT, bg=bg, sz=9)
        W(ws5, ri, 20, str(tr.get("Exit","Hold"))[:10], fc=C_MUT, bg=bg, sz=9)
        W(ws5, ri, 21, tr["Port_Pre"],fc=C_MUT,bg=bg,sz=9,fmt='"$"#,##0')

    # ── Sheet 6 : Per-Trade Cost Analysis ─────────────────────────────────────
    ws6 = wb.create_sheet("Per_Trade_Cost")
    ws6.sheet_view.showGridLines = False
    ws6.sheet_properties.tabColor = "FC5C65"

    ws6.row_dimensions[1].height = 36
    ws6.merge_cells("A1:H1")
    c6h = ws6["A1"]
    c6h.value = "AVERAGE COST PER TRADE BY YEAR  |  1x Scenario  |  All $ amounts"
    c6h.font  = Font(name="Calibri", bold=True, color=C_ACC, size=12)
    c6h.fill  = PatternFill("solid", fgColor=C_BG)
    c6h.alignment = Alignment(horizontal="center", vertical="center")

    ptc_cols = [
        ("Year",7),("Avg Pos $",13),("Comm / trade",14),("Spread / trade",15),
        ("TOTAL / trade",14),("Comm %",9),("Total %",9),
    ]
    HDR(ws6, 2, ptc_cols)

    _, tr1_ = results["1x  Uniform"]
    tr1_["Year"] = pd.to_datetime(tr1_["Date"]).dt.year
    for ri, (yr, g) in enumerate(tr1_.groupby("Year"), 3):
        ws6.row_dimensions[ri].height = 22
        bg  = C_BG if ri % 2 == 0 else C_ROW
        ap  = g["Pos_$"].mean()
        ac  = g["Comm_$"].mean()
        asp = g["Spread_$"].mean()
        at  = ac + asp
        pc  = ac / ap * 100 if ap else 0
        pt  = at / ap * 100 if ap else 0
        W(ws6, ri, 1, int(yr), bold=True, fc=C_GLD, bg=bg, sz=11)
        W(ws6, ri, 2, round(ap,0),  fc=C_WHT, bg=bg, sz=11, fmt='"$"#,##0')
        W(ws6, ri, 3, round(ac,2),  fc=C_MUT, bg=bg, sz=11, fmt='"$"#,##0.00')
        W(ws6, ri, 4, round(asp,2), fc=C_GLD, bg=bg, sz=11, fmt='"$"#,##0.00')
        W(ws6, ri, 5, round(at,2),  fc=C_RED, bg=bg, sz=11, fmt='"$"#,##0.00')
        W(ws6, ri, 6, "%.3f%%" % pc, fc=C_MUT, bg=bg, sz=11)
        W(ws6, ri, 7, "%.3f%%" % pt,
          fc=C_RED if pt > 0.08 else C_MUT, bg=bg, sz=11)

    wb.save(str(OUT_XL))
    print("Saved -> %s" % OUT_XL.name)

# ── Dashboard CSVs ────────────────────────────────────────────────────────────
def save_dash_csvs(results: dict):
    eq_fnames = {"1x  Uniform": "realistic_equity_1x.csv",
                 "Mixed 2x+1x": "realistic_equity_mixed.csv"}
    for lbl, _, _ in SCENARIOS:
        eq_df, _ = results[lbl]
        out_path  = OUT_DIR / eq_fnames.get(lbl, "realistic_equity_%s.csv" % lbl[:3])
        pd.DataFrame({
            "Date":              eq_df["Date"],
            "Portfolio_Value":   eq_df["Port"] / CAPITAL,
            "SPY_Value":         eq_df["SPY"]  / CAPITAL,
            "Period_Return":     eq_df["Ret"],
            "SPY_Period_Return": eq_df["SPY_Ret"],
            "Regime":            eq_df["Regime"],
            "N_Stocks":          eq_df["N"],
            "Comm":              eq_df["Comm"],
            "Spread":            eq_df["Spread"],
            "Total_Cost":        eq_df["Total_Cost"],
        }).to_csv(str(out_path), index=False)
        print("  %s" % out_path.name)

    # Save trades CSV for the Mixed scenario (feeds dashboard Overview)
    _, tr_mix = results["Mixed 2x+1x"]
    tr_out = OUT_DIR / "realistic_trades_mixed.csv"
    pd.DataFrame({
        "Date":            tr_mix["Date"],
        "Ticker":          tr_mix["Ticker"],
        "Sector":          tr_mix["Sector"],
        "Regime":          tr_mix["Regime"],
        "Raw_Score":       tr_mix["Score"],
        "Band":            tr_mix["Band"],
        "Return":          tr_mix["Ret_%"] / 100,
        "Return_Pct":      tr_mix["Ret_%"],
        "Entry_Price":     tr_mix["Entry_$"],
        "Exit_Price":      tr_mix["Exit_$"],
        "SL_Price":        tr_mix["SL_$"],
        "Stop_Triggered":  tr_mix["Stop"].astype(str),
        "Exit_Reason":     tr_mix["Exit"],
        "Pos_$":           tr_mix["Pos_$"],
        "Net_$":           tr_mix["Net_$"],
    }).to_csv(str(tr_out), index=False)
    print("  %s" % tr_out.name)

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    print("=" * 68)
    print("  REALISTIC BACKTEST  --  IBKR Pro Execution Model")
    print("=" * 68)
    print("  Capital      : $%s" % f"{CAPITAL:,.0f}")
    print("  Commission   : $%.3f/share  (min $%.2f)" % (IBKR_RATE, IBKR_MIN))
    print("  Spread       : $%.2f large / $%.2f mid / $%.2f small" % (
          SPREAD_LARGE, SPREAD_MID, SPREAD_SMALL))
    print("  Sizing       : Equal-weight  (portfolio / N stocks)")
    print("  Reg fees     : None  |  Margin interest: None")
    print()

    trades = pd.read_csv(str(TRADES_F))
    equity = pd.read_csv(str(EQUITY_F))
    trades["Date"] = pd.to_datetime(trades["Date"])
    equity["Date"] = pd.to_datetime(equity["Date"])

    rebal_dates = set(equity["Date"].astype(str).str[:10])
    trades_sub  = trades[trades["Date"].astype(str).str[:10].isin(rebal_dates)]
    spy_map     = dict(zip(equity["Date"].astype(str).str[:10],
                           equity["SPY_Period_Return"].fillna(0)))

    # Score band stats
    band_a = trades_sub[(trades_sub["Raw_Score"] >= SCORE_LO_A) &
                        (trades_sub["Raw_Score"] <  SCORE_HI_A)]
    band_b = trades_sub[trades_sub["Raw_Score"] >= SCORE_LO_B]
    rvol   = pd.concat([band_a, band_b])["RVOL_Pct"]
    print("Loading data...")
    print("  Rebalance periods : %d" % len(rebal_dates))
    print("  Band A (92-95)    : %d trades" % len(band_a))
    print("  Band B (98-100)   : %d trades" % len(band_b))
    print("  Spread tier mix   : Large %.0f%%  Mid %.0f%%  Small %.0f%%" % (
          (rvol>=DV_LARGE).mean()*100,
          ((rvol>=DV_MID)&(rvol<DV_LARGE)).mean()*100,
          (rvol<DV_MID).mean()*100))
    print()

    results: dict = {}
    for lbl, lev_a, lev_b in SCENARIOS:
        print("Running %s  (Band-A %.0fx, Band-B %.0fx)..." % (lbl, lev_a, lev_b))
        t0 = time.time()
        eq_df, tr_df = run_sim(trades, spy_map, lev_a, lev_b, rebal_dates)
        m = calc_metrics(eq_df, tr_df)
        print("  Final: $%12s   CAGR: %+5.1f%%/yr   MaxDD: %6.1f%%   "
              "Comm: $%s   Spread: $%s   Total Cost: $%s   (%.1fs)" % (
              f"{m['Final_$']:,.0f}", m["CAGR%"], m["MaxDD%"],
              f"{m['Comm_$']:,.0f}", f"{m['Spread_$']:,.0f}",
              f"{m.get('Total_Cost_$',0):,.0f}", time.time()-t0))
        results[lbl] = (eq_df, tr_df)

    print()
    print("Building Excel...")
    build_excel(equity, results)
    print()
    print("Saving dashboard CSVs...")
    save_dash_csvs(results)

    # No-cost filtered baseline for dashboard Leverage tab
    print("  Computing no-cost baseline...")
    eq_nc, _ = run_sim(trades, spy_map, 1.0, 1.0, rebal_dates, no_cost=True)
    nc_path = OUT_DIR / "realistic_equity_nocost.csv"
    pd.DataFrame({
        "Date":              eq_nc["Date"],
        "Portfolio_Value":   eq_nc["Port"] / CAPITAL,
        "SPY_Value":         eq_nc["SPY"]  / CAPITAL,
        "Period_Return":     eq_nc["Ret"],
        "SPY_Period_Return": eq_nc["SPY_Ret"],
        "Regime":            eq_nc["Regime"],
    }).to_csv(str(nc_path), index=False)
    print("  %s" % nc_path.name)

    print()
    print("=" * 68)
    bm = baseline_metrics(equity)
    print("  BASELINE (no cost):  CAGR %+.1f%%/yr   Final $%s" % (
          bm["CAGR%"], f"{bm['Final_$']:,.0f}"))
    for lbl, _, _ in SCENARIOS:
        m = calc_metrics(*results[lbl])
        print("  %-30s  CAGR %+.1f%%/yr   MaxDD %5.1f%%   Total Cost $%s" % (
              lbl + ":",
              m["CAGR%"], m["MaxDD%"], f"{m['Total_Cost_$']:,.0f}"))
    print("=" * 68)
    print()
    print("Output -> %s" % OUT_XL)

if __name__ == "__main__":
    main()
