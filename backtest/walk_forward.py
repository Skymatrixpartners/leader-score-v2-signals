"""
Walk-Forward Validation
========================
Splits 10-year backtest into:
  In-Sample  (IS ) : 2016-01-01 → 2021-12-31   (train window, 6 years)
  Out-of-Sample (OOS): 2022-01-01 → today       (blind window, ~4.5 years)

Resets capital to $100K for each period so CAGR is directly comparable.
Prints IS vs OOS vs Full-period comparison for both scenarios.

Rule of thumb: if OOS CAGR >= 60% of IS CAGR → real edge confirmed.

Usage:
    python backtest/walk_forward.py
"""
from __future__ import annotations
import numpy  as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT     = Path(__file__).resolve().parents[1]
TRADES_F = ROOT / "output" / "backtest" / "combined_trades_10y.csv"
EQUITY_F = ROOT / "output" / "backtest" / "combined_equity_10y.csv"

# --- Constants (must match realistic_backtest_10y.py) ----------------------
CAPITAL      = 100_000.0
IBKR_RATE    = 0.005;   IBKR_MIN    = 1.00
SPREAD_LARGE = 0.01;    SPREAD_MID  = 0.02;  SPREAD_SMALL = 0.03
DV_LARGE     = 66.0;    DV_MID      = 33.0
SCORE_LO_A   = 92.0;    SCORE_HI_A  = 95.0
SCORE_LO_B   = 98.0

SCENARIOS = [
    ("1x  Uniform", 1.0, 1.0),
    ("Mixed 2x+1x", 2.0, 1.0),
]

IS_END  = "2021-12-31"   # last date of in-sample window
OOS_START = "2022-01-01" # first date of out-of-sample window

# ---------------------------------------------------------------------------
# Cost helpers
# ---------------------------------------------------------------------------
def ibkr_comm(shares: float) -> float:
    return float(max(abs(shares) * IBKR_RATE, IBKR_MIN))

def bid_ask(shares: float, rvol_pct: float) -> float:
    if rvol_pct >= DV_LARGE:   return abs(shares) * SPREAD_LARGE
    if rvol_pct >= DV_MID:     return abs(shares) * SPREAD_MID
    return abs(shares) * SPREAD_SMALL

# Old run_sim (equity only, used for summary metrics)
def run_sim(trades: pd.DataFrame, spy_map: dict,
            lev_a: float, lev_b: float, rebal_dates: set) -> pd.DataFrame:
    eq, _ = run_sim_with_trades(trades, spy_map, lev_a, lev_b, rebal_dates)
    return eq

# ---------------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------------
BG   = "FF0D1117"; HDR_BG = "FF161B2E"; ROW_BG = "FF1A2035"
ACC  = "FF6C63FF"; GRN    = "FF26DE81"; RED    = "FFFC5C65"
GLD  = "FFF0C040"; MUT    = "FF7B82A0"; WHT    = "FFFFFFFF"
IS_C = "FF0D1F3A"; OOS_C  = "FF001F19"

def _bdr():
    s = Side(style="thin", color="FF1F2937")
    return Border(left=s, right=s, top=s, bottom=s)

def W(ws, r, c, v=None, bold=False, fc=WHT, bg=None, hal="center", fmt=None, sz=10):
    cl = ws.cell(row=r, column=c, value=v)
    cl.font      = Font(name="Calibri", bold=bold, color=fc, size=sz)
    cl.alignment = Alignment(horizontal=hal, vertical="center")
    cl.border    = _bdr()
    if bg:   cl.fill = PatternFill("solid", fgColor=bg)
    if fmt:  cl.number_format = fmt
    return cl

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def vc(v): return GRN if (v or 0) > 0 else (RED if (v or 0) < 0 else MUT)

def run_sim_with_trades(trades: pd.DataFrame, spy_map: dict,
                        lev_a: float, lev_b: float,
                        rebal_dates: set) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Returns (equity_df, trades_df) with full per-trade detail."""
    portfolio = CAPITAL
    spy_val   = CAPITAL
    eq_rows   = []
    tr_rows   = []

    for date in sorted(d for d in trades["Date"].unique()
                       if str(d)[:10] in rebal_dates):
        all_grp = trades[trades["Date"] == date]
        band_a  = all_grp[(all_grp["Raw_Score"] >= SCORE_LO_A) &
                          (all_grp["Raw_Score"] <  SCORE_HI_A)]
        band_b  = all_grp[all_grp["Raw_Score"] >= SCORE_LO_B]
        grp     = pd.concat([band_a, band_b]).drop_duplicates()
        if grp.empty:
            continue

        n = len(grp); pv_base = portfolio / n
        gross = comm_t = spread_t = 0.0
        regime = str(grp["Regime"].iloc[0])

        for _, row in grp.iterrows():
            ep   = float(row.get("Entry_Price") or 0)
            if ep <= 0: continue
            sc   = float(row.get("Raw_Score") or 0)
            rvol = float(row.get("RVOL_Pct") or 50)
            ret  = float(row["Return"])
            xp   = float(row.get("Exit_Price") or 0)
            sl   = float(row.get("SL_Price") or 0)
            lev  = lev_a if SCORE_LO_A <= sc < SCORE_HI_A else lev_b
            pv_each = pv_base * lev
            sh   = pv_each / ep
            c_rt = ibkr_comm(sh) * 2
            sp   = bid_ask(sh, rvol)
            gross    += ret * pv_each
            comm_t   += c_rt
            spread_t += sp
            stop_v = str(row.get("Stop_Triggered", "")) in ("True", "true", "1")
            band_lbl = "92-95" if SCORE_LO_A <= sc < SCORE_HI_A else "98-100"
            tr_rows.append({
                "Date":     str(date)[:10],
                "Ticker":   row["Ticker"],
                "Sector":   row.get("Sector", ""),
                "Regime":   regime,
                "Band":     band_lbl,
                "Score":    round(sc, 1),
                "Lev":      lev,
                "Entry_$":  round(ep, 2),
                "Exit_$":   round(xp, 2),
                "SL_$":     round(sl, 2),
                "Shares":   int(round(sh)),
                "Pos_$":    round(pv_each, 2),
                "Ret_%":    round(ret * 100, 2),
                "Gross_$":  round(ret * pv_each, 2),
                "Comm_$":   round(c_rt, 2),
                "Spread_$": round(sp, 2),
                "Net_$":    round(ret * pv_each - c_rt - sp, 2),
                "Stop":     stop_v,
                "Exit":     row.get("Exit_Reason", "Hold"),
                "Port_Pre": round(portfolio, 0),
            })

        total_cost = comm_t + spread_t
        pp = portfolio
        portfolio += gross - total_cost
        spy_ret = float(spy_map.get(str(date)[:10], 0))
        spy_val *= (1 + spy_ret)
        eq_rows.append({
            "Date": str(date)[:10], "Port": round(portfolio, 2),
            "SPY":  round(spy_val, 2),
            "Ret":  (gross - total_cost) / pp if pp > 0 else 0,
            "SPY_Ret": spy_ret, "N": n, "Regime": regime,
        })

    return pd.DataFrame(eq_rows), pd.DataFrame(tr_rows)


def build_excel(out_path: Path,
                is_eq: pd.DataFrame,  is_tr: pd.DataFrame,
                oos_eq: pd.DataFrame, oos_tr: pd.DataFrame,
                year_rows: list, scenario_label: str):
    wb = Workbook(); wb.remove(wb.active)

    def title_row(ws, txt, ncols):
        ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
        c = ws["A1"]
        c.value = txt
        c.font      = Font(name="Calibri", bold=True, color=ACC, size=13)
        c.fill      = PatternFill("solid", fgColor=BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 36

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws = wb.create_sheet("Summary"); ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = "6C63FF"
    title_row(ws, f"WALK-FORWARD VALIDATION  |  {scenario_label}  |  IS: 2016-2021  |  OOS: 2022-present", 8)

    m_is  = calc_metrics(is_eq)
    m_oos = calc_metrics(oos_eq)
    ratio = m_oos["CAGR%"] / m_is["CAGR%"] if m_is["CAGR%"] > 0 else 0
    verdict = ("✓ STRONG (OOS ≥80% of IS)" if ratio >= 0.80 else
               "✓ SOLID  (OOS ≥60% of IS)" if ratio >= 0.60 else
               "~ WEAK   (OOS <60% of IS)" if ratio >= 0.40 else
               "✗ FAIL   (likely overfit)")

    hdr_cols = ["Metric", "IS (2016-2021)", "OOS (2022-now)", "Edge Check"]
    for ci, h in enumerate(hdr_cols, 1):
        W(ws, 2, ci, h, bold=True, fc=ACC, bg=HDR_BG, sz=10)
        ws.row_dimensions[2].height = 24
    set_col_widths(ws, [32, 22, 22, 34])

    rows = [
        ("Period",         m_is["Period"],           m_oos["Period"],          ""),
        ("Years",          f"{m_is['Years']:.1f}",   f"{m_oos['Years']:.1f}",  ""),
        ("CAGR (%/yr)",    f"{m_is['CAGR%']:+.1f}%", f"{m_oos['CAGR%']:+.1f}%", f"OOS/IS = {ratio:.2f}x   {verdict}"),
        ("SPY CAGR",       f"{m_is['SPY_CAGR%']:+.1f}%", f"{m_oos['SPY_CAGR%']:+.1f}%", ""),
        ("Max Drawdown",   f"{m_is['MaxDD%']:.1f}%", f"{m_oos['MaxDD%']:.1f}%",""),
        ("Sharpe",         f"{m_is['Sharpe']:.2f}",  f"{m_oos['Sharpe']:.2f}", ""),
        ("Calmar",         f"{m_is['Calmar']:.2f}",  f"{m_oos['Calmar']:.2f}", ""),
        ("Final $100K",    f"${m_is['Final_$']:,.0f}", f"${m_oos['Final_$']:,.0f}", ""),
        ("Rebal Periods",  str(m_is["Periods"]),     str(m_oos["Periods"]),    ""),
        ("IS Trades",      str(len(is_tr)),           "",                       ""),
        ("OOS Trades",     "",                        str(len(oos_tr)),         ""),
    ]
    for ri, (metric, iv, ov, note) in enumerate(rows, 3):
        ws.row_dimensions[ri].height = 22
        bg = IS_C if ri % 2 == 0 else ROW_BG
        W(ws, ri, 1, metric, fc=WHT, bg=bg, hal="left")
        W(ws, ri, 2, iv,     fc=vc(None) if "+" in str(iv) else WHT, bg=IS_C)
        W(ws, ri, 3, ov,     fc=vc(None) if "+" in str(ov) else WHT, bg=OOS_C)
        W(ws, ri, 4, note,   fc=GLD if "STRONG" in note or "SOLID" in note else (RED if "FAIL" in note else MUT), bg=bg)

    # OOS yearly table
    yr_start = len(rows) + 4
    W(ws, yr_start, 1, "OOS Year-by-Year  (Mixed 2x+1x)", bold=True, fc=ACC, bg=HDR_BG, hal="left")
    ws.merge_cells(f"A{yr_start}:D{yr_start}")
    ws.row_dimensions[yr_start].height = 22
    for ci, h in enumerate(["Year", "Return", "Final $", "Active Days"], 1):
        W(ws, yr_start+1, ci, h, bold=True, fc=ACC, bg=HDR_BG)
    ws.row_dimensions[yr_start+1].height = 20
    for ri2, yr_r in enumerate(year_rows, yr_start+2):
        ws.row_dimensions[ri2].height = 20
        ret_v = yr_r["Return_%"]
        W(ws, ri2, 1, yr_r["Year"],        fc=WHT, bg=ROW_BG)
        W(ws, ri2, 2, f"{ret_v:+.1f}%",   fc=GRN if ret_v > 0 else RED, bg=ROW_BG)
        W(ws, ri2, 3, f"${yr_r['Final_$']:,.0f}", fc=WHT, bg=ROW_BG)
        W(ws, ri2, 4, yr_r["Active_Days"], fc=WHT, bg=ROW_BG)

    def write_trades_sheet(ws, tr_df, period_label, tab_color, bg_period):
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = tab_color
        ws.freeze_panes = "A3"
        ncols = 20
        title_row(ws, f"{period_label}  |  {len(tr_df)} trades  |  {scenario_label}", ncols)
        cols = [
            ("Date",8),("Ticker",8),("Sector",14),("Regime",14),("Band",8),
            ("Score",7),("Lev",5),("Entry $",9),("Exit $",9),("SL $",9),
            ("Shares",8),("Pos $",11),("Ret %",8),("Gross $",11),
            ("Comm $",9),("Spread $",9),("Net $",11),("Stop",7),("Exit",10),("Port Pre $",12),
        ]
        for ci, (h, w) in enumerate(cols, 1):
            W(ws, 2, ci, h, bold=True, fc=ACC, bg=HDR_BG, sz=9)
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[2].height = 24

        col_keys = ["Date","Ticker","Sector","Regime","Band","Score","Lev",
                    "Entry_$","Exit_$","SL_$","Shares","Pos_$","Ret_%",
                    "Gross_$","Comm_$","Spread_$","Net_$","Stop","Exit","Port_Pre"]
        for ri, (_, row) in enumerate(tr_df.iterrows(), 3):
            ws.row_dimensions[ri].height = 16
            bg = bg_period if ri % 2 == 0 else ROW_BG
            for ci, key in enumerate(col_keys, 1):
                val = row.get(key, "")
                fc  = WHT
                if key == "Ret_%":   fc = GRN if float(val or 0) > 0 else RED
                if key == "Net_$":   fc = GRN if float(val or 0) > 0 else RED
                if key == "Stop":    fc = RED if val is True else WHT
                W(ws, ri, ci, val, fc=fc, bg=bg, sz=9)

    ws_is  = wb.create_sheet("IS Trades (2016-2021)")
    ws_oos = wb.create_sheet("OOS Trades (2022-now)")
    write_trades_sheet(ws_is,  is_tr,  "IN-SAMPLE  2016–2021",    "3A86FF", IS_C)
    write_trades_sheet(ws_oos, oos_tr, "OUT-OF-SAMPLE  2022–now", "00D4AA", OOS_C)

    # Equity sheet
    ws_eq = wb.create_sheet("Equity Curves"); ws_eq.sheet_view.showGridLines = False
    ws_eq.sheet_properties.tabColor = "F0C040"
    title_row(ws_eq, f"Equity Curves  |  {scenario_label}", 6)
    for ci, h in enumerate(["Period","Date","Portfolio $","SPY $","Period Ret %","SPY Ret %"], 1):
        W(ws_eq, 2, ci, h, bold=True, fc=ACC, bg=HDR_BG)
    set_col_widths(ws_eq, [12, 12, 14, 14, 14, 14])
    ws_eq.row_dimensions[2].height = 24; ri = 3
    for period_lbl, eq_df, bg in [("IS", is_eq, IS_C), ("OOS", oos_eq, OOS_C)]:
        for _, row in eq_df.iterrows():
            ws_eq.row_dimensions[ri].height = 15
            rb = bg if ri % 2 == 0 else ROW_BG
            W(ws_eq, ri, 1, period_lbl, fc=ACC, bg=rb)
            W(ws_eq, ri, 2, row["Date"], fc=WHT, bg=rb)
            W(ws_eq, ri, 3, f"${row['Port']:,.0f}", fc=WHT, bg=rb)
            W(ws_eq, ri, 4, f"${row['SPY']:,.0f}", fc=WHT, bg=rb)
            ret_pct = round(row["Ret"] * 100, 2)
            W(ws_eq, ri, 5, f"{ret_pct:+.2f}%", fc=GRN if ret_pct > 0 else RED, bg=rb)
            spy_pct = round(row["SPY_Ret"] * 100, 2)
            W(ws_eq, ri, 6, f"{spy_pct:+.2f}%", fc=GRN if spy_pct > 0 else RED, bg=rb)
            ri += 1

    wb.save(str(out_path))
    print(f"  Saved: {out_path.name}")


    return float(max(abs(shares) * IBKR_RATE, IBKR_MIN))

def bid_ask(shares: float, rvol_pct: float) -> float:
    if rvol_pct >= DV_LARGE:   return abs(shares) * SPREAD_LARGE
    if rvol_pct >= DV_MID:     return abs(shares) * SPREAD_MID
    return abs(shares) * SPREAD_SMALL

def run_sim(trades: pd.DataFrame, spy_map: dict,
            lev_a: float, lev_b: float, rebal_dates: set) -> pd.DataFrame:
    """Returns equity DataFrame (one row per rebalance date)."""
    portfolio = CAPITAL
    spy_val   = CAPITAL
    eq_rows   = []

    for date in sorted(d for d in trades["Date"].unique()
                       if str(d)[:10] in rebal_dates):
        all_grp = trades[trades["Date"] == date]
        band_a  = all_grp[(all_grp["Raw_Score"] >= SCORE_LO_A) &
                          (all_grp["Raw_Score"] <  SCORE_HI_A)]
        band_b  = all_grp[all_grp["Raw_Score"] >= SCORE_LO_B]
        grp     = pd.concat([band_a, band_b]).drop_duplicates()
        if grp.empty:
            continue

        n = len(grp); pv_base = portfolio / n
        gross = comm_t = spread_t = 0.0

        for _, row in grp.iterrows():
            ep   = float(row.get("Entry_Price") or 0)
            if ep <= 0: continue
            sc   = float(row.get("Raw_Score") or 0)
            rvol = float(row.get("RVOL_Pct") or 50)
            ret  = float(row["Return"])
            lev  = lev_a if SCORE_LO_A <= sc < SCORE_HI_A else lev_b
            sh   = (pv_base * lev) / ep
            gross    += ret * pv_base * lev
            comm_t   += ibkr_comm(sh) * 2
            spread_t += bid_ask(sh, rvol)

        total_cost = comm_t + spread_t
        pp         = portfolio
        portfolio += gross - total_cost
        spy_ret    = float(spy_map.get(str(date)[:10], 0))
        spy_val   *= (1 + spy_ret)

        eq_rows.append({
            "Date":    str(date)[:10],
            "Port":    round(portfolio, 2),
            "SPY":     round(spy_val, 2),
            "Ret":     (gross - total_cost) / pp if pp > 0 else 0,
            "SPY_Ret": spy_ret,
            "N":       n,
        })

    return pd.DataFrame(eq_rows)

def calc_metrics(eq: pd.DataFrame, label: str = "") -> dict:
    if eq.empty:
        return {"Label": label, "CAGR%": 0, "MaxDD%": 0, "Sharpe": 0,
                "Calmar": 0, "Trades": 0, "Final_$": CAPITAL,
                "SPY_CAGR%": 0}
    pv   = eq["Port"].values
    sv   = eq["SPY"].values
    rets = eq["Ret"].dropna().values
    d0   = pd.to_datetime(eq["Date"].iloc[0])
    d1   = pd.to_datetime(eq["Date"].iloc[-1])
    ny   = max((d1 - d0).days / 365.25, 0.1)

    cagr   = ((pv[-1] / CAPITAL) ** (1 / ny) - 1) * 100
    s_cagr = ((sv[-1] / CAPITAL) ** (1 / ny) - 1) * 100

    pk = CAPITAL; mdd = 0.0
    for v in pv:
        pk = max(pk, v); mdd = min(mdd, (v - pk) / pk)

    sharpe = (np.mean(rets) / np.std(rets) * np.sqrt(52)
              if np.std(rets) > 0 else 0)
    calmar = cagr / abs(mdd * 100) if mdd < 0 else 0

    return {
        "Label":     label,
        "Period":    f"{d0.date()} → {d1.date()}",
        "Years":     round(ny, 1),
        "CAGR%":     round(cagr, 1),
        "SPY_CAGR%": round(s_cagr, 1),
        "MaxDD%":    round(mdd * 100, 1),
        "Sharpe":    round(sharpe, 2),
        "Calmar":    round(calmar, 2),
        "Final_$":   round(pv[-1]),
        "Periods":   len(eq),
    }

# ---------------------------------------------------------------------------
def main():
    print("\n" + "=" * 70)
    print("  WALK-FORWARD VALIDATION  |  Leader Score V2")
    print("  In-Sample : 2016 – 2021  |  Out-of-Sample: 2022 – present")
    print("=" * 70)

    trades = pd.read_csv(str(TRADES_F))
    equity = pd.read_csv(str(EQUITY_F))
    trades["Date"] = pd.to_datetime(trades["Date"])
    equity["Date"] = pd.to_datetime(equity["Date"])

    spy_map = dict(zip(
        equity["Date"].astype(str).str[:10],
        equity["SPY_Period_Return"].fillna(0)
    ))
    all_dates = set(equity["Date"].astype(str).str[:10])

    is_dates  = {d for d in all_dates if d <= IS_END}
    oos_dates = {d for d in all_dates if d >= OOS_START}

    print(f"\n  Rebalance periods → IS: {len(is_dates)}  |  OOS: {len(oos_dates)}  |  Total: {len(all_dates)}")

    # Count trades per period
    t_dates = trades["Date"].astype(str).str[:10]
    is_t  = trades[(t_dates <= IS_END)]
    oos_t = trades[(t_dates >= OOS_START)]

    ba_is  = is_t[(is_t["Raw_Score"] >= SCORE_LO_A) & (is_t["Raw_Score"] < SCORE_HI_A)]
    bb_is  = is_t[is_t["Raw_Score"] >= SCORE_LO_B]
    ba_oos = oos_t[(oos_t["Raw_Score"] >= SCORE_LO_A) & (oos_t["Raw_Score"] < SCORE_HI_A)]
    bb_oos = oos_t[oos_t["Raw_Score"] >= SCORE_LO_B]

    print(f"  IS  trades → Band A: {len(ba_is)}  Band B: {len(bb_is)}  Total: {len(ba_is)+len(bb_is)}")
    print(f"  OOS trades → Band A: {len(ba_oos)}  Band B: {len(bb_oos)}  Total: {len(ba_oos)+len(bb_oos)}")

    # Run simulations
    for lbl, lev_a, lev_b in SCENARIOS:
        print(f"\n{'─'*70}")
        print(f"  SCENARIO: {lbl}  (Band-A {lev_a:.0f}x / Band-B {lev_b:.0f}x)")
        print(f"{'─'*70}")

        eq_is   = run_sim(trades, spy_map, lev_a, lev_b, is_dates)
        eq_oos  = run_sim(trades, spy_map, lev_a, lev_b, oos_dates)
        eq_full = run_sim(trades, spy_map, lev_a, lev_b, all_dates)

        m_is   = calc_metrics(eq_is,   "In-Sample  (2016-2021)")
        m_oos  = calc_metrics(eq_oos,  "Out-of-Sample (2022-now)")
        m_full = calc_metrics(eq_full, "Full Period (2016-now)")

        # Header
        print(f"\n  {'Metric':<28} {'IS (2016-21)':>18} {'OOS (2022+)':>18} {'Full':>18}")
        print(f"  {'':28} {'─'*18} {'─'*18} {'─'*18}")

        rows = [
            ("Period",      m_is['Period'],   m_oos['Period'],   m_full['Period']),
            ("Years",       f"{m_is['Years']:.1f}",   f"{m_oos['Years']:.1f}",   f"{m_full['Years']:.1f}"),
            ("CAGR (%/yr)", f"{m_is['CAGR%']:+.1f}%", f"{m_oos['CAGR%']:+.1f}%", f"{m_full['CAGR%']:+.1f}%"),
            ("SPY CAGR",    f"{m_is['SPY_CAGR%']:+.1f}%", f"{m_oos['SPY_CAGR%']:+.1f}%", f"{m_full['SPY_CAGR%']:+.1f}%"),
            ("Max Drawdown",f"{m_is['MaxDD%']:.1f}%", f"{m_oos['MaxDD%']:.1f}%", f"{m_full['MaxDD%']:.1f}%"),
            ("Sharpe",      f"{m_is['Sharpe']:.2f}",  f"{m_oos['Sharpe']:.2f}",  f"{m_full['Sharpe']:.2f}"),
            ("Calmar",      f"{m_is['Calmar']:.2f}",  f"{m_oos['Calmar']:.2f}",  f"{m_full['Calmar']:.2f}"),
            ("Final $100K →",f"${m_is['Final_$']:>12,.0f}", f"${m_oos['Final_$']:>12,.0f}", f"${m_full['Final_$']:>12,.0f}"),
        ]
        for metric, is_v, oos_v, full_v in rows:
            print(f"  {metric:<28} {is_v:>18} {oos_v:>18} {full_v:>18}")

        # Edge check
        oos_ratio = m_oos["CAGR%"] / m_is["CAGR%"] if m_is["CAGR%"] > 0 else 0
        print()
        print(f"  Edge Check: OOS CAGR / IS CAGR = {oos_ratio:.2f}x", end="  ")
        if   oos_ratio >= 0.80: print("✓ STRONG  (OOS within 20% of IS)")
        elif oos_ratio >= 0.60: print("✓ SOLID   (OOS within 40% of IS)")
        elif oos_ratio >= 0.40: print("~ WEAK    (OOS degraded >40%)")
        else:                   print("✗ FAIL    (OOS degraded >60% — likely overfit)")

    # Per-year breakdown for OOS
    print(f"\n{'─'*70}")
    print("  OOS Per-Year Returns  (Mixed 2x+1x)")
    print(f"{'─'*70}")
    eq_oos_mx = run_sim(trades, spy_map, 2.0, 1.0, oos_dates)
    if not eq_oos_mx.empty:
        eq_oos_mx["Year"] = pd.to_datetime(eq_oos_mx["Date"]).dt.year
        prev = CAPITAL
        print(f"  {'Year':<8} {'Return':>10} {'Final $':>15} {'Active Days':>12}")
        for yr, g in eq_oos_mx.groupby("Year"):
            end = float(g["Port"].iloc[-1])
            ret = (end / prev - 1) * 100
            print(f"  {yr:<8} {ret:>+9.1f}%  ${end:>13,.0f}  {len(g):>10}")
            prev = end

    # Save results to CSV + Excel
    out_dir = ROOT / "output" / "backtest"
    out_dir.mkdir(parents=True, exist_ok=True)

    summary_rows = []
    for lbl, lev_a, lev_b in SCENARIOS:
        eq_is   = run_sim(trades, spy_map, lev_a, lev_b, is_dates)
        eq_oos  = run_sim(trades, spy_map, lev_a, lev_b, oos_dates)
        eq_full = run_sim(trades, spy_map, lev_a, lev_b, all_dates)
        for period, eq in [("IS_2016-2021", eq_is), ("OOS_2022-now", eq_oos), ("Full_2016-now", eq_full)]:
            m = calc_metrics(eq)
            summary_rows.append({
                "Scenario":  lbl, "Period": period,
                "Date_Range":m.get("Period",""), "Years": m["Years"],
                "CAGR_%":    m["CAGR%"],  "SPY_CAGR_%": m["SPY_CAGR%"],
                "MaxDD_%":   m["MaxDD%"], "Sharpe": m["Sharpe"],
                "Calmar":    m["Calmar"], "Final_$": m["Final_$"],
                "Periods":   m["Periods"],
            })

    # OOS per-year (Mixed 2x+1x)
    eq_oos_yr = run_sim(trades, spy_map, 2.0, 1.0, oos_dates)
    year_rows = []
    if not eq_oos_yr.empty:
        eq_oos_yr["Year"] = pd.to_datetime(eq_oos_yr["Date"]).dt.year
        prev = CAPITAL
        for yr, g in eq_oos_yr.groupby("Year"):
            end = float(g["Port"].iloc[-1])
            year_rows.append({"Year": yr, "Return_%": round((end/prev-1)*100,1),
                              "Final_$": round(end), "Active_Days": len(g)})
            prev = end

    pd.DataFrame(summary_rows).to_csv(str(out_dir/"walk_forward_summary.csv"), index=False)
    pd.DataFrame(year_rows).to_csv(str(out_dir/"walk_forward_oos_yearly.csv"), index=False)
    print(f"  Saved: walk_forward_summary.csv")
    print(f"  Saved: walk_forward_oos_yearly.csv")

    # Build Excel for Mixed 2x+1x (the primary scenario)
    print("\nBuilding Excel...")
    lev_a, lev_b = 2.0, 1.0
    lbl = "Mixed 2x+1x"
    is_eq_xl,  is_tr_xl  = run_sim_with_trades(trades, spy_map, lev_a, lev_b, is_dates)
    oos_eq_xl, oos_tr_xl = run_sim_with_trades(trades, spy_map, lev_a, lev_b, oos_dates)
    xl_path = out_dir / "walk_forward_10y.xlsx"
    build_excel(xl_path, is_eq_xl, is_tr_xl, oos_eq_xl, oos_tr_xl, year_rows, lbl)

    print(f"  Folder: {out_dir}")
    print(f"\n{'='*70}\n")

if __name__ == "__main__":
    main()
